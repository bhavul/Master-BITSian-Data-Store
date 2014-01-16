import openpyxl as xl
from openpyxl import Workbook
from openpyxl.cell import get_column_letter
import os
import re
import sqlite3 as sql
import Levenshtein
import smtplib
import sys


def addToDatabase(con, dataset, priority, date):

	db = con.cursor()

	query = "SELECT column_ID, name FROM columns"
	db.execute(query)
	results = db.fetchall()

	keyColumns = dict()
	for result in results:
		keyColumns[result['name']] = result['column_ID']


	query = "SELECT max(ID)+1 ID FROM master"
	db.execute(query)
	result = db.fetchone()
	masterID = result['ID'] if result['ID'] else 1

	query = "SELECT max(update_ID)+1 ID FROM updates"
	db.execute(query)
	result = db.fetchone()
	updateID = result['ID'] if result['ID'] else 1


	lst = list()

	for da in dataset:
		for key,value in da.items():
			columnValue = keyColumns[key]
			lst.append((masterID,columnValue,value,updateID))

		masterID += 1

	query = "INSERT INTO master VALUES (?,?,?,?)" 
	db.executemany(query,lst)

	query = "INSERT INTO updates VALUES (%d, datetime('now'), %d, '%s', 0)" % (updateID, priority, date)
	db.execute(query)

	con.commit()
	db.close()



def compare(newData, dataset, weights, threshold=0.8):

	result = list()

	while len(newData)>0:
		flag = False

		newRecord = newData.pop(0)

		for i in xrange(len(dataset)):
			record = dataset[i]
			
			match = compareRecords(newRecord, record, weights)

			if match>=threshold:
				record = dataset.pop(i)

				for attr in newRecord:
					if attr=='ID':
						continue

					if attr not in record:
						record[attr] = list()

					record[attr].extend(newRecord[attr])

				result.append((record['ID'], newRecord['ID']))
				newData.append(record)
				flag = True
				break

		if not flag:
			dataset.append(newRecord)

	return result


def compareRecords(r1, r2, weights):

	similarity = 0.0

	for attr, l1 in r1.items():
		
		if attr not in r2 or attr=='ID':
			continue

		l2 = r2[attr]
		t, w = weights[attr]

		match = list()

		for i1 in l1:
			for i2 in l2:
				try:
					match.append(Levenshtein.ratio(i1.upper(), i2.upper()))
				except:
					pass

		match.sort(reverse=True)

		if match and match[0]>=t:
				similarity += match[0] * w

	return similarity


def createExcelFile(con, filepath, dataset):
	db = con.cursor()

	query = "SELECT column_ID, name from columns"
	db.execute(query)
	results = db.fetchall()

	columns = dict()
	for result in results:
		columns[result['column_ID']] = result['name']

	wb = Workbook()
	ws = wb.worksheets[0]
	
	for col_id, col_name in columns.items():
		ws.cell(get_column_letter(col_id)+'1').value = col_name

	row = 1
	for data in dataset:
		row += 1
		for key, value in data.items():
			if key in columns:
				ws.cell(get_column_letter(key)+str(row)).value = "; ".join(value)

	wb.save(filename = filepath+'.xlsx')
	db.close()


def createRecords(datadump, single=False):

	dataset = dict()

	for data in datadump:
		ID = data['ID']
		attr = data['column_ID']
		
		if ID not in dataset:
			dataset[ID] = dict()
			dataset[ID]['ID'] = ID

		if attr not in dataset[ID]:
			dataset[ID][attr] = list()

		if not single or len(dataset[ID][attr])==0:
			dataset[ID][attr].append(data['value'])

	return dataset.values()


def createLabels(con, excel, row, csv, sheet_no=0):
	
	table = xl.load_workbook(excel)
	sheets = table.get_sheet_names()
	sheet = table.get_sheet_by_name(sheets[sheet_no])

	db = con.cursor()
	query = "SELECT name FROM columns"
	db.execute(query)
	columns = db.fetchall()

	attr = list()

	col = 0
	flag = False

	while True:
		col += 1

		value = str(sheet.cell(get_column_letter(col) + row).value).upper()

		if value=='NONE':
			if flag==True:
				break
			else:
				flag = True

		else:
			flag = False
		

		column_name = dict()
		for column in columns:
			name = column['name'].encode('ascii', 'ignore')
			ratio =  Levenshtein.ratio(value, name.upper())

			if ratio>0.5 and value!='NONE':
				column_name[ratio] = name

		if column_name.keys():
			m = max(column_name.keys(), key=float)
			attr.append(column_name[m])

		else:
			attr.append("")
			

	with open(csv, 'w') as fp:
		fp.write(",".join(attr))


def getDataset(bitsaa_columns, path, columns, start, sheet_no=0):

	table = xl.load_workbook(path)
	sheets = table.get_sheet_names()

	column_map = dict()
	column_index = 1

	for column in columns:
		if column in bitsaa_columns:
			column_map[column] = get_column_letter(column_index)

		column_index += 1

	sheet = table.get_sheet_by_name(sheets[sheet_no])

	dataset = list()

	for row in range(start, len(sheet.rows)+1):
		data = dict()

		for column in columns:
			try:
				value = str(sheet.cell(column_map[column] + str(row)).value)

				if value!='None':
					match = re.search(bitsaa_columns[column], value)

					if match.group(0):
						data[column] = value
			except:
				pass

		dataset.append(data)

	return dataset


def getDist(con, filepath):

	db = con.cursor()

	results = list()

	query = "SELECT m.ID, m.column_ID, m.value FROM master m NATURAL JOIN columns c NATURAL JOIN updates u WHERE c.type=0 ORDER BY m.ID, m.column_ID, u.priority, u.date DESC, u.updated_ON DESC"
	db.execute(query)
	results.extend(db.fetchall())

	query = "SELECT m.ID, m.column_ID, m.value FROM master m NATURAL JOIN columns c NATURAL JOIN updates u WHERE c.type=1 ORDER BY m.ID, m.column_ID, u.date DESC, u.updated_ON DESC"
	db.execute(query)
	results.extend(db.fetchall())

	dataset = createRecords(results, True)

	db.close()

	createExcelFile(con, filepath, dataset)


def getMaster(con, filepath):

	db = con.cursor()

	query = "SELECT m.* FROM master m NATURAL JOIN updates u ORDER BY m.ID DESC, m.column_ID ASC"
	db.execute(query)
	dataset = createRecords(db.fetchall())


	db.close()

	createExcelFile(con, filepath, dataset)


def getOriginalFiles(con, directory):
	db = con.cursor()

	if not os.path.exists(directory):
		os.mkdir(directory)

	query = "SELECT * FROM master"
	db.execute(query)
	results = db.fetchall()
	db.close()

	datadump = dict()
	for result in results:
		update_ID = result['update_ID']

		if update_ID not in datadump:
			datadump[update_ID] = list()

		datadump[update_ID].append(result)


	for update_ID, temp in datadump.items():
		filepath = "%s/Data%d" % (directory, update_ID)
		dataset = createRecords(temp)
		createExcelFile(con, filepath, dataset)
	

def maintainance(con):

	db = con.cursor()

	query = "SELECT m.* FROM master m NATURAL JOIN updates u WHERE u.status=0 ORDER BY m.ID DESC"
	db.execute(query)
	newData = createRecords(db.fetchall())

	query = "SELECT m.* FROM master m NATURAL JOIN updates u WHERE u.status=1 ORDER BY m.ID DESC"
	db.execute(query)
	dataset = createRecords(db.fetchall())

	query = "SELECT column_ID, threshold, weight FROM columns"
	db.execute(query)
	results = db.fetchall()

	weights = dict()
	for result in results:
		weights[result['column_ID']] = (result['threshold'], result['weight'])


	updates = compare(newData, dataset, weights)

	query = "UPDATE master SET ID=? WHERE ID=?"
	db.executemany(query, updates)

	query = "UPDATE updates SET STATUS=1"
	db.execute(query)

	con.commit()
	db.close()


def sendemail(from_addr, to_addr_list, cc_addr_list,
              subject, message,
              login, password,
              smtpserver):
    header  = 'From: %s\n' % from_addr
    header += 'To: %s\n' % ','.join(to_addr_list)
    header += 'Cc: %s\n' % ','.join(cc_addr_list)
    header += 'Subject: %s\n\n' % subject
    message = header + message
  
    server = smtplib.SMTP(smtpserver)
    server.starttls()
    server.login(login,password)
    problems = server.sendmail(from_addr, to_addr_list, message)
    server.quit()

    return problems


def setup(path, db_name='bitsaa.db'):
	con = sql.connect(db_name)
	db = con.cursor()

	query = "CREATE TABLE master (ID INTEGER, column_ID INTEGER, value TEXT, update_ID INTEGER)"
	db.execute(query)

	query = "CREATE TABLE updates (update_ID INTEGER, updated_on TIMESTAMP, priority INTEGER, date DATE, status INTEGER)"
	db.execute(query)

	query = "CREATE TABLE columns(column_ID INTEGER, name TEXT, type INTEGER, threshold REAL(3,2), weight REAL(3,2), regex TEXT)"
	db.execute(query)


	with open(path, 'r') as fp:
		lst = list()
		lines = fp.readlines()


	lst = list()

	for line in lines:
		match = line[:-1].split(",")
		lst.append(tuple(match))

	query = "INSERT INTO columns VALUES (?, ?, ?, ?, ?, ?)"
	db.executemany(query, lst)

	con.commit()
	db.close()
	con.close()


def main():
	try:
		option = sys.argv[1]

		if option=='-s' or option=='--setup':
			setup(sys.argv[2])

		else:

			con = sql.connect('bitsaa.db')
			con.row_factory = sql.Row

			if option=='-i' or option=='--insert':

				if len(sys.argv)>=7:
					columns_path = sys.argv[6]
				else:
					columns_path = sys.argv[2]+'.csv'

				with open(columns_path, 'r') as fp:
					line = fp.readline()

				columns = line.split(",")


				db = con.cursor()
				db.execute("SELECT name, regex FROM columns")

				bitsaa_columns = dict()
				for result in db.fetchall():
					bitsaa_columns[result['name']] = result['regex']


				dataset = getDataset(bitsaa_columns, sys.argv[2], columns, int(sys.argv[5]))
				addToDatabase(con, dataset, int(sys.argv[3]), sys.argv[4])


			elif option=='-l' or option=='--labels':
				if len(sys.argv)>=5:
					createLabels(con, sys.argv[2], sys.argv[3], sys.argv[4])
				else:
					createLabels(con, sys.argv[2], sys.argv[3], sys.argv[2]+'.csv')

			elif option=='-m' or option=='--maintain':
				maintainance(con)
			
			elif option=='-o' or option=='--original':
				getOriginalFiles(con, sys.argv[2])

			elif option=='-p' or option=='-purge':
				try:
					getOriginalFiles(con, "Backup")
				except:
					pass

				con.close()
				os.remove('bitsaa.db')

				adminList = []
				ccList = []
				sendemail(
					from_addr    = 'bitsaamail@gmail.com', 
					to_addr_list = adminList,
					cc_addr_list = ccList, 
					subject      = 'Privileged Action Performed on BITSAA Database', 
					message      = 'This is to inform you that an admin recently wiped the master bitsaa database.', 
					login        = 'bitsaamail', 
					password     = 'randompassword',
					smtpserver   = 'smtp.gmail.com:587')

			elif option=='-xd' or option=='--dist':
				getDist(con, sys.argv[2])

			elif option=='-xm' or option=='--master':
				getMaster(con, sys.argv[2])

	except:
		print '''
Usage: python %s OPTION FILE|DIRECTORY [PRIORITY DATE START] [FILE]

OPTIONS:
\t-i, --insert\tInsert data from excel sheet into the database
\t-l, --labels\tTries to predict the appropriate labels
\t-m, --maintain\tGroups similar records and removes duplicates
\t-o, --original\tGet back the original Excel sheets
\t-p, --purge\tDeletes the database
\t-s, --setup\tPerfroms initial setup (to be run only once)
\t-xd, --dist\tGet excel sheet of the database (for distribution)
\t-xm, --master\tGet excel sheet of the master database (for admin)
''' % sys.argv[0]


if __name__ == "__main__":
	main()